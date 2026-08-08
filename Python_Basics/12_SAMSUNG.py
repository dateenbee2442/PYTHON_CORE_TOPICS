import openpyxl as qq
wb = qq.load_workbook('SAMSUNG_corrected.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(2, 4)
print(cell.value)


# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     nextt = cell.value - 25000
#     corrected_price = nextt // 210
#     corrected_price_cell = sheet.cell(row, 4)
#     corrected_price_cell.value = corrected_price
# wb.save('SAMSUNG_corrected.xlsx')