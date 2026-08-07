import openpyxl as px
current = px.load_workbook('iphonexlsx.xlsx')
sheet = current['Sheet1']
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = (cell.value - 25000) // 210
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
current.save('iphone2.xlsx')