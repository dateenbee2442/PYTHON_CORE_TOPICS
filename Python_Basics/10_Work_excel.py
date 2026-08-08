import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def apply_discount_and_create_chart(filename):
    # 1. Load the file
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # 2. Apply 10% discount and put in Column D
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price  

    # 3. Set up chart data from Column D  
    values = Reference(sheet, 
                    min_row=2, 
                    max_row=sheet.max_row, 
                    min_col=4, 
                    max_col=4)

    # 4. Create and add chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # 5. Save new file
    wb.save(filename)